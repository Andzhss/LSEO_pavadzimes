import streamlit as st
import datetime
import pandas as pd
import json
import os
import io
import requests
import base64

# --- Google Bibliotēkas ---
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload

from utils import scrape_lursoft, money_to_words_lv
from pdf_generator import generate_pdf
from docx_generator import generate_docx

# --- Konfigurācija ---
st.set_page_config(page_title="SIA Baltic SEO Invoice Generator", layout="wide")

# --- Sky blue theme CSS ---
st.markdown("""
<style>
    /* Main background */
    .stApp {
        background: linear-gradient(135deg, #f0f7ff 0%, #e8f4fd 100%);
    }
    /* Sidebar fons */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1a6fa8 0%, #1557a0 100%);
    }
    /* Visi teksti sānjoslā — balti */
    section[data-testid="stSidebar"] * {
        color: #ffffff !important;
    }
    /* Ievadlauku foni — balti ar tumšu tekstu */
    section[data-testid="stSidebar"] input,
    section[data-testid="stSidebar"] textarea,
    section[data-testid="stSidebar"] [data-baseweb="input"] input,
    section[data-testid="stSidebar"] [data-baseweb="select"] div,
    section[data-testid="stSidebar"] [data-baseweb="select"] span,
    section[data-testid="stSidebar"] [role="listbox"] * {
        background-color: #ffffff !important;
        color: #1a1a2e !important;
    }
    /* Selectbox un number input konteineri */
    section[data-testid="stSidebar"] [data-baseweb="select"] > div,
    section[data-testid="stSidebar"] [data-baseweb="base-input"] {
        background-color: #ffffff !important;
        border-color: #4da6e8 !important;
    }
    section[data-testid="stSidebar"] [data-baseweb="base-input"] input {
        color: #1a1a2e !important;
    }
    /* Date input */
    section[data-testid="stSidebar"] [data-baseweb="input"] {
        background-color: #ffffff !important;
    }
    /* Etiķetes virs laukiem — gaiši zilas */
    section[data-testid="stSidebar"] label,
    section[data-testid="stSidebar"] .stSelectbox label,
    section[data-testid="stSidebar"] .stNumberInput label,
    section[data-testid="stSidebar"] .stDateInput label,
    section[data-testid="stSidebar"] .stTextInput label {
        color: #cce8ff !important;
        font-size: 0.85rem;
        font-weight: 600;
    }
    /* Headers */
    h1 { color: #1a6fa8 !important; border-bottom: 3px solid #4da6e8; padding-bottom: 8px; }
    h2, h3 { color: #1557a0 !important; }
    /* Buttons */
    .stButton > button {
        background: linear-gradient(90deg, #1a6fa8, #4da6e8);
        color: white;
        border: none;
        border-radius: 6px;
        font-weight: 600;
    }
    .stButton > button:hover {
        background: linear-gradient(90deg, #1557a0, #1a6fa8);
        transform: translateY(-1px);
        box-shadow: 0 4px 12px rgba(26,111,168,0.3);
    }
    /* Download buttons */
    .stDownloadButton > button {
        background: linear-gradient(90deg, #0e7c3a, #1aab52) !important;
        color: white !important;
        border: none !important;
        border-radius: 6px !important;
        font-weight: 600 !important;
    }
    /* Cards / expanders */
    .streamlit-expanderHeader {
        background: #d0eaf8 !important;
        border-radius: 6px !important;
        color: #1a6fa8 !important;
        font-weight: 600 !important;
    }
    /* Info boxes */
    .stInfo { background: #d0eaf8 !important; border-left: 4px solid #1a6fa8 !important; }
    /* Dividers */
    hr { border-color: #4da6e8 !important; }
    /* Tabs */
    .stTabs [data-baseweb="tab"] {
        background: #d0eaf8;
        border-radius: 6px 6px 0 0;
        color: #1a6fa8;
        font-weight: 600;
    }
    .stTabs [aria-selected="true"] {
        background: #1a6fa8 !important;
        color: white !important;
    }
</style>
""", unsafe_allow_html=True)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CREDENTIALS_FILE = os.path.join(BASE_DIR, "credentials.json")
TOKEN_FILE = os.path.join(BASE_DIR, "token.json")

# Lokālie CSV faili
LOCAL_PRESETS_PATH    = os.path.join(BASE_DIR, "presets.csv")
LOCAL_HISTORY_PATH    = os.path.join(BASE_DIR, "invoice_history.csv")
LOCAL_TEST_HIST_PATH  = os.path.join(BASE_DIR, "test_invoice_history.csv")

# GitHub ceļi (repozitorijā)
GITHUB_REPO            = "Andzhss/OnlinePavadzimes"
GITHUB_PRESETS_PATH    = "OnlinePavadzimes/presets.csv"
GITHUB_HISTORY_PATH    = "OnlinePavadzimes/invoice_history.csv"
GITHUB_TEST_HIST_PATH  = "OnlinePavadzimes/test_invoice_history.csv"

# Google Drive
GOOGLE_DRIVE_FOLDER_ID = "1vqhkHGH9WAMaFnXtduyyjYdEzHMx0iX9"
SCOPES = ['https://www.googleapis.com/auth/drive.file']

# CSV kolonnas vēsturei
HISTORY_COLS = [
    'kartas_nr', 'datums', 'pr_partneris', 'pr_pvn_nr',
    'pr_datums', 'pr_numurs', 'darijuma_apraksts',
    'vertiba_bez_pvn', 'dabas_resursi', 'atlaides', 'pvn_summa', 'kopeja_summa',
    'due_date', 'client_reg_no', 'client_address', 'doc_type', 'items_json', 'comments', 'created_at'
]

# ---------------------------------------------------------------------------
# Valodas tulkošana
# ---------------------------------------------------------------------------

TRANSLATIONS = {
    "lv": {
        "app_title": "SIA Baltic SEO Rēķinu Ģenerators",
        "tab_invoice": "📄 Rēķina izveide",
        "tab_presets": "⚙️ Produktu sagataves",
        "settings": "Rēķina iestatījumi",
        "load_github": "☁️ Ielādēt vēsturi no GitHub",
        "history_updated": "Vēsture atjaunota!",
        "history_sync_fail": "Neizdevās sinhronizēt (tukša vēsture vai nav Token)",
        "doc_nr": "Dokumenta Nr.",
        "doc_id_label": "**Dokumenta ID:**",
        "last_invoice": "📋 Pēdējā pavadzīme:",
        "date": "Datums",
        "due_date": "Apmaksāt līdz",
        "doc_type": "Dokumenta tips",
        "open_prev": "📂 Atvērt iepriekšējo pavadzīmi",
        "choose_doc": "Izvēlies dokumentu",
        "load_selected": "📂 Ielādēt izvēlēto",
        "no_saved": "Nav saglabātu pavadzīmju.",
        "test_load": "🔄 Testa pavadzīmju ielāde",
        "choose_test": "Izvēlies testa dokumentu",
        "load_test": "Ielādēt izvēlēto",
        "no_test": "Nav saglabātu testa pavadzīmju.",
        "google_drive": "Google Drive",
        "open_drive": "📂 Atvērt Google Drive mapi",
        "connected": "✅ Pieslēgts",
        "disconnect": "Atslēgties",
        "not_connected": "❌ Nav pieslēgts",
        "data_mgmt": "Datu pārvaldība",
        "delete_history": "🗑️ Dzēst visu rēķinu vēsturi",
        "confirm_delete": "Vai tiešām dzēst visu vēsturi? Nevar atsaukt.",
        "yes_delete": "Jā, dzēst",
        "cancel": "Atcelt",
        "client": "Klients",
        "lursoft_url": "Lursoft saite",
        "load_lursoft": "Ielādēt datus no Lursoft",
        "loading": "Datu ielasīšana...",
        "lursoft_ok": "Dati veiksmīgi ielasīti!",
        "lursoft_fail": "Neizdevās ielasīt datus.",
        "name": "Nosaukums",
        "address": "Adrese",
        "reg_no": "Reģ. Nr.",
        "vat_no": "PVN Nr.",
        "receiver": "Saņēmējs",
        "receiver_inst": "Iestādes nosaukums (Saņēmējs)",
        "receiver_reg": "Reģistrācijas numurs (Saņēmējs)",
        "receiver_addr": "Juridiskā adrese (Saņēmējs)",
        "customer": "Pasūtītājs",
        "customer_name": "Nosaukums (Pasūtītājs)",
        "customer_reg": "Reģistrācijas numurs (Pasūtītājs)",
        "customer_addr": "Juridiskā adrese (Pasūtītājs)",
        "goods": "Preces / Pakalpojumi",
        "add_from_presets": "Pievienot no sagatavēm",
        "choose_product": "Izvēlieties produktu",
        "quantity": "Daudzums",
        "add_to_table": "➕ Pievienot tabulai",
        "presets_empty": "Sagatavju saraksts ir tukšs. Pievienojiet tos cilnē 'Produktu sagataves'.",
        "recalculate": "🔄 Pārrēķināt summas",
        "discount": "### Atlaide",
        "discount_type": "Atlaides veids:",
        "no_discount": "Nav atlaides",
        "percent_discount": "Procentos (%)",
        "eur_discount": "Ciparos (EUR)",
        "discount_pct": "Atlaides procenti (%)",
        "discount_eur_input": "Atlaides summa (EUR)",
        "advance_settings": "### Avansa iestatījumi",
        "calc_method": "Aprēķina veids:",
        "in_eur": "Ciparos (EUR)",
        "in_pct": "Procentos (%)",
        "total_order": "Kopējā pasūtījuma summa:",
        "advance_payable": "APMAKSĀJAMAIS AVANSS",
        "amount_words_label": "**Summa vārdiem (Avanss):**",
        "total_no_vat": "**KOPĀ (bez PVN un atlaides):**",
        "discount_amount": "**Atlaides apjoms",
        "total_with_discount": "**Kopā ar atlaidi (bez PVN):**",
        "vat_21": "**PVN (21%):**",
        "vat_toggle": "Pievienot PVN (21%)",
        "total_payable": "**KOPUMĀ APMAKSAI:**",
        "amount_words_full": "**Summa vārdiem:**",
        "calc_error": "Kļūda aprēķinos:",
        "comments_header": "Komentāri un Paraksti",
        "comments_label": "Papildus komentāri / piezīmes (tiks iekļauti dokumentā)",
        "prepared_by": "Dokumentu sagatavoja",
        "title": "Amats",
        "default_title": "valdes loceklis",
        "signatory_preview": "Paraksta laukā būs:",
        "download_header": "### Lejupielāde un Arhivēšana",
        "proforma_toggle": "📝 Ģenerēt kā Proformas (testa) dokumentu",
        "proforma_help": "Ja ieslēgts: dokuments sauksies 'Proformas...', saglabāsies testa vēsturē un NETIKS augšupielādēts Google Drive.",
        "download_pdf": "📄 Lejupielādēt PDF",
        "download_word": "📝 Lejupielādēt Word",
        "pdf_error": "Kļūda PDF:",
        "word_error": "Kļūda Word:",
        "history_expander": "🗄️ Rēķinu vēsture (Izrakstītie)",
        "download_excel": "📥 Lejupielādēt kā Excel",
        "history_empty": "Vēsture ir tukša.",
        "presets_header": "Produktu un Pakalpojumu Sagataves",
        "presets_desc": "Šeit varat pievienot, labot un dzēst biežāk izmantotos produktus.",
        "no_token_warning": "⚠️ GitHub Token nav atrasts. Izmaiņas tiks saglabātas tikai lokāli un pēc servera restartēšanas pazudīs.",
        "import_github": "⬇️ Importēt no GitHub (Atjaunot)",
        "import_ok": "Sagataves veiksmīgi ielādētas no GitHub!",
        "import_fail": "Neizdevās ielādēt sagataves no GitHub (pārbaudiet Token un faila esamību).",
        "save_presets": "💾 Saglabāt izmaiņas sagatavēs",
        "doc_types": ["Pavadzīme", "Rēķins", "Avansa rēķins", "E-rēķins"],
        "proforma_map": {
            "Pavadzīme": "Proformas pavadzīme",
            "Rēķins": "Proformas rēķins",
            "Avansa rēķins": "Proformas avansa rēķins"
        },
        "save_github_ok": "✅ Proformas dokuments saglabāts vēsturē (GitHub)",
        "save_ok": "✅ Dokuments saglabāts vēsturē (GitHub)",
        "save_fail": "⚠️ Kļūda saglabājot GitHub:",
        "drive_ok": "✅ Saglabāts Drive:",
        "drive_fail": "⚠️ Kļūda saglabājot Drive",
        "drive_no_conn": "Nav pieslēgts Google Drive (tikai lejupielādēts)",
        "auth_step1": "**[1. Klikšķini šeit, lai autorizētos Google]",
        "auth_step2": "2. Iekopē kodu šeit:",
        "auth_step3": "3. Apstiprināt kodu",
        "auth_ok": "Veiksmīgi pieslēgts!",
        "auth_fail": "Kļūda:",
        "auth_no_code": "Lūdzu ievadi kodu!",
        "no_credentials": "Trūkst credentials.json faila!",
        "saved_local_github": "Saglabāts lokāli un",
        "saved_local_only": "Saglabāts lokāli! (Nav GitHub Token)",
        "github_token_how": "Kā pieslēgt GitHub Token?",
        "language": "🌐 Valoda / Language",
    },
    "en": {
        "app_title": "SIA Baltic SEO Invoice Generator",
        "tab_invoice": "📄 Create Invoice",
        "tab_presets": "⚙️ Product Presets",
        "settings": "Invoice Settings",
        "load_github": "☁️ Load history from GitHub",
        "history_updated": "History updated!",
        "history_sync_fail": "Sync failed (empty history or no Token)",
        "doc_nr": "Document No.",
        "doc_id_label": "**Document ID:**",
        "last_invoice": "📋 Last invoice:",
        "date": "Date",
        "due_date": "Due date",
        "doc_type": "Document type",
        "open_prev": "📂 Open previous invoice",
        "choose_doc": "Choose document",
        "load_selected": "📂 Load selected",
        "no_saved": "No saved invoices.",
        "test_load": "🔄 Load test invoices",
        "choose_test": "Choose test document",
        "load_test": "Load selected",
        "no_test": "No saved test invoices.",
        "google_drive": "Google Drive",
        "open_drive": "📂 Open Google Drive folder",
        "connected": "✅ Connected",
        "disconnect": "Disconnect",
        "not_connected": "❌ Not connected",
        "data_mgmt": "Data management",
        "delete_history": "🗑️ Delete all invoice history",
        "confirm_delete": "Really delete all history? Cannot be undone.",
        "yes_delete": "Yes, delete",
        "cancel": "Cancel",
        "client": "Client",
        "lursoft_url": "Lursoft URL",
        "load_lursoft": "Load data from Lursoft",
        "loading": "Loading data...",
        "lursoft_ok": "Data loaded successfully!",
        "lursoft_fail": "Failed to load data.",
        "name": "Name",
        "address": "Address",
        "reg_no": "Reg. No.",
        "vat_no": "VAT No.",
        "receiver": "Receiver",
        "receiver_inst": "Institution name (Receiver)",
        "receiver_reg": "Registration number (Receiver)",
        "receiver_addr": "Legal address (Receiver)",
        "customer": "Customer",
        "customer_name": "Name (Customer)",
        "customer_reg": "Registration number (Customer)",
        "customer_addr": "Legal address (Customer)",
        "goods": "Goods / Services",
        "add_from_presets": "Add from presets",
        "choose_product": "Choose product",
        "quantity": "Quantity",
        "add_to_table": "➕ Add to table",
        "presets_empty": "Preset list is empty. Add them in the 'Product presets' tab.",
        "recalculate": "🔄 Recalculate totals",
        "discount": "### Discount",
        "discount_type": "Discount type:",
        "no_discount": "No discount",
        "percent_discount": "Percentage (%)",
        "eur_discount": "Amount (EUR)",
        "discount_pct": "Discount percentage (%)",
        "discount_eur_input": "Discount amount (EUR)",
        "advance_settings": "### Advance settings",
        "calc_method": "Calculation method:",
        "in_eur": "Amount (EUR)",
        "in_pct": "Percentage (%)",
        "total_order": "Total order amount:",
        "advance_payable": "ADVANCE PAYABLE",
        "amount_words_label": "**Amount in words (Advance):**",
        "total_no_vat": "**TOTAL (excl. VAT and discount):**",
        "discount_amount": "**Discount amount",
        "total_with_discount": "**Total with discount (excl. VAT):**",
        "vat_21": "**VAT (21%):**",
        "vat_toggle": "Apply VAT (21%)",
        "total_payable": "**TOTAL PAYABLE:**",
        "amount_words_full": "**Amount in words:**",
        "calc_error": "Calculation error:",
        "comments_header": "Comments and Signatures",
        "comments_label": "Additional comments / notes (will be included in document)",
        "prepared_by": "Document prepared by",
        "title": "Title",
        "default_title": "board member",
        "signatory_preview": "Signature field will show:",
        "download_header": "### Download and Archive",
        "proforma_toggle": "📝 Generate as Proforma (test) document",
        "proforma_help": "If enabled: document will be named 'Proforma...', saved in test history and will NOT be uploaded to Google Drive.",
        "download_pdf": "📄 Download PDF",
        "download_word": "📝 Download Word",
        "pdf_error": "PDF error:",
        "word_error": "Word error:",
        "history_expander": "🗄️ Invoice history (Issued)",
        "download_excel": "📥 Download as Excel",
        "history_empty": "History is empty.",
        "presets_header": "Product and Service Presets",
        "presets_desc": "Here you can add, edit and delete frequently used products.",
        "no_token_warning": "⚠️ GitHub Token not found. Changes will only be saved locally and will be lost after server restart.",
        "import_github": "⬇️ Import from GitHub (Refresh)",
        "import_ok": "Presets successfully loaded from GitHub!",
        "import_fail": "Failed to load presets from GitHub (check Token and file existence).",
        "save_presets": "💾 Save preset changes",
        "doc_types": ["Invoice", "Receipt", "Advance invoice", "E-invoice"],
        "proforma_map": {
            "Invoice": "Proforma invoice",
            "Receipt": "Proforma receipt",
            "Advance invoice": "Proforma advance invoice"
        },
        "save_github_ok": "✅ Proforma document saved to history (GitHub)",
        "save_ok": "✅ Document saved to history (GitHub)",
        "save_fail": "⚠️ Error saving to GitHub:",
        "drive_ok": "✅ Saved to Drive:",
        "drive_fail": "⚠️ Error saving to Drive",
        "drive_no_conn": "Google Drive not connected (download only)",
        "auth_step1": "**[1. Click here to authorize Google]",
        "auth_step2": "2. Paste code here:",
        "auth_step3": "3. Confirm code",
        "auth_ok": "Successfully connected!",
        "auth_fail": "Error:",
        "auth_no_code": "Please enter the code!",
        "no_credentials": "Missing credentials.json file!",
        "saved_local_github": "Saved locally and",
        "saved_local_only": "Saved locally! (No GitHub Token)",
        "github_token_how": "How to connect GitHub Token?",
        "language": "🌐 Language / Valoda",
    }
}

def t(key):
    lang = st.session_state.get("lang", "lv")
    return TRANSLATIONS[lang].get(key, key)

# ---------------------------------------------------------------------------
# GitHub palīgfunkcijas
# ---------------------------------------------------------------------------

def get_github_token():
    token = st.secrets.get("GITHUB_TOKEN", "")
    return token.strip().strip('"').strip("'") if token else ""

def fetch_csv_from_github(github_path):
    token = get_github_token()
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{github_path}"
    headers = {"Accept": "application/vnd.github.v3+json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    try:
        r = requests.get(url, headers=headers, timeout=10)
        if r.status_code == 200:
            content_b64 = r.json().get("content", "")
            if content_b64:
                return base64.b64decode(content_b64).decode("utf-8")
    except Exception:
        pass
    return None

def push_csv_to_github(df, github_path, commit_message="Update CSV via App"):
    token = get_github_token()
    if not token:
        return False, "Nav GitHub Token"
    try:
        url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{github_path}"
        headers = {
            "Authorization": f"token {token}",
            "Accept": "application/vnd.github.v3+json"
        }
        r = requests.get(url, headers=headers, timeout=10)
        sha = r.json().get("sha", "") if r.status_code == 200 else ""
        csv_content = df.to_csv(index=False)
        encoded = base64.b64encode(csv_content.encode("utf-8")).decode("utf-8")
        data = {"message": commit_message, "content": encoded, "branch": "main"}
        if sha:
            data["sha"] = sha
        put_r = requests.put(url, headers=headers, json=data, timeout=15)
        if put_r.status_code in [200, 201]:
            return True, "Veiksmīgi saglabāts GitHub!" if st.session_state.get("lang","lv")=="lv" else "Successfully saved to GitHub!"
        else:
            return False, f"GitHub error ({put_r.status_code})"
    except Exception as e:
        return False, str(e)

# ---------------------------------------------------------------------------
# Google Drive funkcijas
# ---------------------------------------------------------------------------

def get_drive_service():
    creds = None
    if os.path.exists(TOKEN_FILE):
        try:
            creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
        except Exception:
            os.remove(TOKEN_FILE)
            creds = None
    if creds and creds.expired and creds.refresh_token:
        try:
            creds.refresh(Request())
            with open(TOKEN_FILE, 'w') as f:
                f.write(creds.to_json())
        except Exception:
            if os.path.exists(TOKEN_FILE):
                os.remove(TOKEN_FILE)
            creds = None
    if creds and creds.valid:
        return build('drive', 'v3', credentials=creds)
    return None

def upload_to_drive(file_buffer, filename, mime_type):
    try:
        service = get_drive_service()
        if not service:
            return False
        file_metadata = {'name': filename, 'parents': [GOOGLE_DRIVE_FOLDER_ID]}
        file_buffer.seek(0)
        media = MediaIoBaseUpload(file_buffer, mimetype=mime_type, resumable=True)
        service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        file_buffer.seek(0)
        return True
    except Exception as e:
        st.error(f"❌ Kļūda Google Drive: {e}")
        return False

# ---------------------------------------------------------------------------
# Vēstures funkcijas
# ---------------------------------------------------------------------------

def _fmt(val):
    try:
        return f"{float(val):,.2f}".replace(",", "X").replace(".", ",").replace("X", " ")
    except Exception:
        return str(val)

def _migrate_old_history(df):
    records = []
    for i, (_, row) in enumerate(df.iterrows(), 1):
        items_str = row.get('items', '[]')
        try:
            items_list = json.loads(items_str) if pd.notna(items_str) and items_str else []
        except Exception:
            items_list = []
        base = sum(
            float(it.get('raw_qty', 0) or 0) * float(it.get('raw_price', 0) or 0)
            for it in items_list
        )
        try:
            total_val = float(str(row.get('total', '0')).replace('\u00a0', '').replace(' ', '').replace(',', '.'))
        except Exception:
            total_val = 0.0
        vat = round(total_val - base, 2)
        descriptions = [it.get('name', '') for it in items_list if it.get('name')]
        rec = {
            'kartas_nr':         i,
            'datums':            row.get('date', ''),
            'pr_partneris':      row.get('client_name', ''),
            'pr_pvn_nr':         row.get('client_vat_no', row.get('client_reg_no', '')),
            'pr_datums':         row.get('date', ''),
            'pr_numurs':         row.get('doc_id', ''),
            'darijuma_apraksts': '; '.join(descriptions),
            'vertiba_bez_pvn':   _fmt(base),
            'dabas_resursi':     '',
            'atlaides':          _fmt(0),
            'pvn_summa':         _fmt(vat),
            'kopeja_summa':      row.get('total', ''),
            'due_date':          row.get('due_date', ''),
            'client_reg_no':     row.get('client_reg_no', ''),
            'client_address':    row.get('client_address', ''),
            'doc_type':          row.get('doc_type', ''),
            'items_json':        items_str if pd.notna(items_str) else '[]',
            'comments':          row.get('comments', ''),
            'created_at':        row.get('created_at', ''),
            'items':             items_list,
            'doc_id':            row.get('doc_id', ''),
            'client_name':       row.get('client_name', ''),
            'client_vat_no':     row.get('client_vat_no', ''),
            'date':              row.get('date', ''),
            'total':             row.get('total', ''),
        }
        records.append(rec)
    return records

def load_history(local_path):
    if not os.path.exists(local_path):
        return []
    try:
        df = pd.read_csv(local_path, dtype=str)
        if df.empty:
            return []
        if 'doc_id' in df.columns and 'kartas_nr' not in df.columns:
            return _migrate_old_history(df)
        records = []
        for _, row in df.iterrows():
            rec = row.to_dict()
            items_str = rec.get('items_json', '[]')
            try:
                rec['items'] = json.loads(items_str) if pd.notna(items_str) and items_str else []
            except Exception:
                rec['items'] = []
            rec['doc_id']        = rec.get('pr_numurs', '')
            rec['client_name']   = rec.get('pr_partneris', '')
            rec['client_vat_no'] = rec.get('pr_pvn_nr', '')
            rec['date']          = rec.get('datums', '')
            rec['total']         = rec.get('kopeja_summa', '')
            records.append(rec)
        return records
    except Exception:
        return []

def _history_to_df(history):
    rows = []
    for entry in history:
        row = {col: entry.get(col, '') for col in HISTORY_COLS}
        items_val = entry.get('items_json', entry.get('items', []))
        if isinstance(items_val, list):
            row['items_json'] = json.dumps(items_val, ensure_ascii=False)
        else:
            row['items_json'] = items_val if items_val else '[]'
        rows.append(row)
    return pd.DataFrame(rows, columns=HISTORY_COLS) if rows else pd.DataFrame(columns=HISTORY_COLS)

def save_to_history(invoice_data, local_path, github_path):
    history = load_history(local_path)
    items        = invoice_data.get('items', [])
    raw_total    = float(invoice_data.get('raw_total', 0) or 0)
    raw_discount = float(invoice_data.get('raw_discount_eur', 0) or 0)
    apply_vat    = invoice_data.get('apply_vat', True)
    if apply_vat:
        base_amount = round(raw_total / 1.21, 2)
    else:
        base_amount = raw_total
    vat_amount   = round(raw_total - base_amount, 2)
    descriptions = [it.get('name', '') for it in items if it.get('name')]
    pr_numurs    = invoice_data.get('doc_id', '')
    existing_nums = [int(str(e.get('kartas_nr', 0)).strip() or 0) for e in history]
    next_kartas   = max(existing_nums, default=0) + 1
    new_entry = {
        'kartas_nr':         next_kartas,
        'datums':            invoice_data.get('date', ''),
        'pr_partneris':      invoice_data.get('client_name', ''),
        'pr_pvn_nr':         invoice_data.get('client_vat_no', ''),
        'pr_datums':         invoice_data.get('date', ''),
        'pr_numurs':         pr_numurs,
        'darijuma_apraksts': '; '.join(descriptions),
        'vertiba_bez_pvn':   _fmt(base_amount),
        'dabas_resursi':     '',
        'atlaides':          _fmt(raw_discount),
        'pvn_summa':         _fmt(vat_amount),
        'kopeja_summa':      invoice_data.get('total', ''),
        'due_date':          invoice_data.get('due_date', ''),
        'client_reg_no':     invoice_data.get('client_reg_no', ''),
        'client_address':    invoice_data.get('client_address', ''),
        'doc_type':          invoice_data.get('doc_type', ''),
        'items_json':        json.dumps(items, ensure_ascii=False),
        'comments':          invoice_data.get('comments', ''),
        'created_at':        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'items':             items,
        'doc_id':            pr_numurs,
        'client_name':       invoice_data.get('client_name', ''),
        'client_vat_no':     invoice_data.get('client_vat_no', ''),
        'date':              invoice_data.get('date', ''),
        'total':             invoice_data.get('total', ''),
    }
    updated = False
    for i, entry in enumerate(history):
        if entry.get('pr_numurs') == pr_numurs or entry.get('doc_id') == pr_numurs:
            new_entry['kartas_nr'] = entry.get('kartas_nr', next_kartas)
            history[i] = new_entry
            updated = True
            break
    if not updated:
        history.append(new_entry)
    df = _history_to_df(history)
    df.to_csv(local_path, index=False, encoding='utf-8')
    if get_github_token():
        success, msg = push_csv_to_github(df, github_path, f"Pievieno {pr_numurs}")
        return success, msg
    else:
        return False, "Nav GITHUB_TOKEN"

def sync_history_from_github(local_path, github_path):
    content = fetch_csv_from_github(github_path)
    if content and ('doc_id' in content or 'kartas_nr' in content):
        with open(local_path, 'w', encoding='utf-8') as f:
            f.write(content)
        return True
    return False

def get_next_invoice_number(history):
    if not history:
        return 1
    max_num = 0
    for entry in history:
        doc = str(entry.get('pr_numurs', entry.get('doc_id', '')))
        parts = doc.split()
        if len(parts) > 1 and parts[-1].isdigit():
            num = int(parts[-1])
            if num > max_num:
                max_num = num
    return max_num + 1

# ---------------------------------------------------------------------------
# Pavadzīmes ielāde formā
# ---------------------------------------------------------------------------

def load_invoice_into_form(entry):
    doc_id = str(entry.get('pr_numurs', entry.get('doc_id', '')))
    parts = doc_id.split()
    if len(parts) > 1 and parts[-1].isdigit():
        st.session_state.doc_number_input = int(parts[-1])
    st.session_state.client_data = {
        'name':    entry.get('pr_partneris', entry.get('client_name', '')),
        'address': entry.get('client_address', ''),
        'reg_no':  entry.get('client_reg_no', ''),
        'vat_no':  entry.get('pr_pvn_nr', entry.get('client_vat_no', ''))
    }
    items_raw = entry.get('items', [])
    items_list = []
    for item in items_raw:
        try:
            qty   = float(item.get('raw_qty', item.get('qty', 0)) or 0)
            price = float(item.get('raw_price', 0) or 0)
        except Exception:
            qty, price = 0.0, 0.0
        items_list.append({
            "NOSAUKUMS":   item.get('name', ''),
            "Mērvienība":  item.get('unit', ''),
            "DAUDZUMS":    qty,
            "CENA (EUR)":  price
        })
    if items_list:
        st.session_state.items_df = pd.DataFrame(items_list)
    loaded_type = entry.get('doc_type', t("doc_types")[0])
    # Normalize proforma types back to base types
    proforma_reverse = {v: k for k, v in TRANSLATIONS["lv"]["proforma_map"].items()}
    proforma_reverse.update({v: k for k, v in TRANSLATIONS["en"]["proforma_map"].items()})
    loaded_type = proforma_reverse.get(loaded_type, loaded_type)
    st.session_state.loaded_doc_type = loaded_type
    st.session_state.loaded_comments = entry.get('comments', '')
    try:
        date_str = entry.get('datums', entry.get('date', ''))
        if date_str:
            st.session_state.loaded_doc_date = datetime.datetime.strptime(date_str, "%d.%m.%Y").date()
        due_str = entry.get('due_date', '')
        if due_str:
            st.session_state.loaded_due_date = datetime.datetime.strptime(due_str, "%d.%m.%Y").date()
    except Exception:
        pass

# ---------------------------------------------------------------------------
# Lejupielādes callback
# ---------------------------------------------------------------------------

def handle_download(invoice_data, file_buffer, filename, mime_type, is_proforma):
    if is_proforma:
        success, msg = save_to_history(invoice_data, LOCAL_TEST_HIST_PATH, GITHUB_TEST_HIST_PATH)
        if success:
            st.toast(t("save_github_ok"), icon="💾")
        else:
            st.error(f"{t('save_fail')} {msg}")
    else:
        success, msg = save_to_history(invoice_data, LOCAL_HISTORY_PATH, GITHUB_HISTORY_PATH)
        if success:
            st.toast(t("save_ok"), icon="💾")
        else:
            st.error(f"{t('save_fail')} {msg}")

        if get_drive_service():
            success_drive = upload_to_drive(file_buffer, filename, mime_type)
            if success_drive:
                st.toast(f"{t('drive_ok')} {filename}", icon="☁️")
            else:
                st.toast(t("drive_fail"), icon="❌")
        else:
            st.toast(t("drive_no_conn"), icon="⚠️")

# ---------------------------------------------------------------------------
# Sagataves
# ---------------------------------------------------------------------------

def load_presets():
    default_df = pd.DataFrame(columns=["NOSAUKUMS", "Mērvienība", "CENA (EUR)"])
    if os.path.exists(LOCAL_PRESETS_PATH):
        try:
            df = pd.read_csv(LOCAL_PRESETS_PATH)
            if df.empty or "NOSAUKUMS" not in df.columns:
                return default_df
            return df
        except Exception:
            return default_df
    return default_df

def save_presets(df):
    df.to_csv(LOCAL_PRESETS_PATH, index=False)

# ---------------------------------------------------------------------------
# Excel ģenerators vēsturei
# ---------------------------------------------------------------------------

def generate_history_excel(history):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Rēķinu vēsture"

    header_fill  = PatternFill("solid", fgColor="1A6FA8")
    subhead_fill = PatternFill("solid", fgColor="4DA6E8")
    even_fill    = PatternFill("solid", fgColor="EAF4FB")
    odd_fill     = PatternFill("solid", fgColor="FFFFFF")
    header_font  = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
    data_font    = Font(name="Calibri", size=9)
    thin         = Side(style="thin", color="4DA6E8")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers_row1 = [
        ("Kārtas\nNr.",                                                           1, "center"),
        ("Datums",                                                                1, "center"),
        ("PR norādītais\ndarījuma partneris",                                     1, "left"),
        ("PR norādītā darījuma partnera\nreģistrācijas vai PVN maksātāja Nr.",    1, "center"),
        ("PR datums un numurs",                                                   2, "center"),
        ("Darījuma apraksts",                                                     1, "left"),
        ("PR norādītā\ndarījuma vērtība\n(bez PVN)",                              1, "right"),
        ("Dabas resursu\nun akcīzes\nnodokļi",                                    1, "right"),
        ("Piešķirtās\natlaides",                                                  1, "right"),
        ("PVN\nsumma",                                                            1, "right"),
        ("Kopējā\nsumma",                                                         1, "right"),
    ]

    col = 1
    for text, span, align in headers_row1:
        cell = ws.cell(row=1, column=col, value=text)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.border    = border
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if span == 2:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
            ws.cell(row=1, column=col + 1).fill   = header_fill
            ws.cell(row=1, column=col + 1).border = border
        else:
            ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        col += span

    pr_col = 5
    for sub_col, sub_text in [(pr_col, "Datums"), (pr_col + 1, "Numurs")]:
        cell = ws.cell(row=2, column=sub_col, value=sub_text)
        cell.fill      = subhead_fill
        cell.font      = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
        cell.border    = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 18

    for r_idx, entry in enumerate(history, start=3):
        row_fill = even_fill if r_idx % 2 == 0 else odd_fill
        values = [
            (entry.get('kartas_nr', ''),                                        "center"),
            (entry.get('datums', entry.get('date', '')),                        "center"),
            (entry.get('pr_partneris', entry.get('client_name', '')),           "left"),
            (entry.get('pr_pvn_nr', entry.get('client_vat_no', '')),            "center"),
            (entry.get('pr_datums', entry.get('date', '')),                     "center"),
            (entry.get('pr_numurs', entry.get('doc_id', '')),                   "center"),
            (entry.get('darijuma_apraksts', ''),                                "left"),
            (entry.get('vertiba_bez_pvn', ''),                                  "right"),
            (entry.get('dabas_resursi', '') or '—',                             "right"),
            (entry.get('atlaides', '') or '—',                                  "right"),
            (entry.get('pvn_summa', ''),                                        "right"),
            (entry.get('kopeja_summa', entry.get('total', '')),                 "right"),
        ]
        for c_idx, (val, align) in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=str(val) if val else '')
            cell.fill      = row_fill
            cell.font      = data_font
            cell.border    = border
            cell.alignment = Alignment(
                horizontal=align, vertical="center", wrap_text=(align == "left")
            )
        ws.row_dimensions[r_idx].height = 15

    col_widths = [8, 11, 28, 22, 11, 11, 40, 14, 12, 12, 12, 13]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ---------------------------------------------------------------------------
# render_presets_app
# ---------------------------------------------------------------------------

def render_presets_app():
    st.header(t("presets_header"))
    st.write(t("presets_desc"))

    github_token = get_github_token()
    if not github_token:
        st.warning(t("no_token_warning"))
        with st.expander(t("github_token_how")):
            st.markdown("1. Streamlit Cloud → Manage app → Settings → Secrets")
            st.code('GITHUB_TOKEN = "ghp_Jusu_GitHub_Token"')

    if "preset_editor_key" not in st.session_state:
        st.session_state.preset_editor_key = 0

    if st.button(t("import_github")):
        content = fetch_csv_from_github(GITHUB_PRESETS_PATH)
        if content and "NOSAUKUMS" in content and "CENA (EUR)" in content:
            with open(LOCAL_PRESETS_PATH, "w", encoding='utf-8') as f:
                f.write(content)
            st.success(t("import_ok"))
            st.session_state.preset_editor_key += 1
            st.rerun()
        else:
            st.error(t("import_fail"))

    presets_df = load_presets()
    edited_presets = st.data_editor(
        presets_df,
        key=f"presets_editor_{st.session_state.preset_editor_key}",
        num_rows="dynamic",
        width="stretch",
        column_config={"CENA (EUR)": st.column_config.NumberColumn(format="%.2f", step=0.01)}
    )

    if st.button(t("save_presets")):
        save_presets(edited_presets)
        if github_token:
            with st.spinner("Saglabā GitHub..."):
                success, msg = push_csv_to_github(edited_presets, GITHUB_PRESETS_PATH, "Update presets.csv via App")
                if success:
                    st.success(f"{t('saved_local_github')} {msg}")
                else:
                    st.warning(f"Lokāli saglabāts, bet GitHub kļūda: {msg}")
        else:
            st.success(t("saved_local_only"))

# ---------------------------------------------------------------------------
# render_invoice_app
# ---------------------------------------------------------------------------

def render_invoice_app():
    history      = load_history(LOCAL_HISTORY_PATH)
    test_history = load_history(LOCAL_TEST_HIST_PATH)
    next_number  = get_next_invoice_number(history)

    st.sidebar.header(t("settings"))

    # Language toggle
    lang_choice = st.sidebar.radio(
        t("language"),
        ["🇱🇻 Latviešu", "🇬🇧 English"],
        horizontal=True,
        key="lang_radio"
    )
    st.session_state.lang = "lv" if "Latviešu" in lang_choice else "en"

    st.sidebar.markdown("---")

    if st.sidebar.button(t("load_github")):
        ok1 = sync_history_from_github(LOCAL_HISTORY_PATH, GITHUB_HISTORY_PATH)
        ok2 = sync_history_from_github(LOCAL_TEST_HIST_PATH, GITHUB_TEST_HIST_PATH)
        if ok1 or ok2:
            st.sidebar.success(t("history_updated"))
            st.rerun()
        else:
            st.sidebar.warning(t("history_sync_fail"))

    if 'doc_number_input' not in st.session_state:
        st.session_state.doc_number_input = next_number

    doc_number_input = st.sidebar.number_input(
        t("doc_nr"), min_value=1, value=st.session_state.doc_number_input, step=1
    )
    doc_id = f"LSEO {doc_number_input:04d}"
    st.sidebar.markdown(f"{t('doc_id_label')} {doc_id}")

    if history:
        last_num = get_next_invoice_number(history) - 1
        st.sidebar.info(f"{t('last_invoice')} **LSEO {last_num:04d}**")

    default_doc_date = st.session_state.get('loaded_doc_date', datetime.date.today())
    doc_date = st.sidebar.date_input(t("date"), default_doc_date)

    default_due_date = st.session_state.get('loaded_due_date', doc_date + datetime.timedelta(days=14))
    due_date = st.sidebar.date_input(t("due_date"), default_due_date)

    doc_types = t("doc_types")
    dt_index = 0
    if 'loaded_doc_type' in st.session_state:
        ldt = st.session_state.loaded_doc_type
        # Try to match by position
        lv_types = TRANSLATIONS["lv"]["doc_types"]
        en_types = TRANSLATIONS["en"]["doc_types"]
        if ldt in lv_types:
            dt_index = lv_types.index(ldt)
        elif ldt in en_types:
            dt_index = en_types.index(ldt)
        if dt_index < len(doc_types):
            pass
        else:
            dt_index = 0
    doc_type = st.sidebar.selectbox(t("doc_type"), doc_types, index=dt_index)

    st.sidebar.markdown("---")

    st.sidebar.subheader(t("open_prev"))
    if history:
        hist_options = {
            f"{e.get('pr_numurs', e.get('doc_id',''))} — {e.get('pr_partneris', e.get('client_name',''))} ({e.get('datums', e.get('date',''))})": e
            for e in reversed(history)
        }
        selected_hist_label = st.sidebar.selectbox(t("choose_doc"), list(hist_options.keys()), key="hist_select")
        if st.sidebar.button(t("load_selected"), key="load_hist_btn"):
            load_invoice_into_form(hist_options[selected_hist_label])
            st.rerun()
    else:
        st.sidebar.info(t("no_saved"))

    st.sidebar.markdown("---")

    st.sidebar.subheader(t("test_load"))
    if test_history:
        test_options = {
            f"{t_.get('pr_numurs', t_.get('doc_id',''))} — {t_.get('pr_partneris', t_.get('client_name',''))} ({t_.get('datums', t_.get('date',''))})": t_
            for t_ in reversed(test_history)
        }
        selected_test_label = st.sidebar.selectbox(t("choose_test"), list(test_options.keys()))
        if st.sidebar.button(t("load_test"), key="load_test_btn"):
            load_invoice_into_form(test_options[selected_test_label])
            st.rerun()
    else:
        st.sidebar.info(t("no_test"))

    st.sidebar.markdown("---")

    st.sidebar.subheader(t("google_drive"))
    if GOOGLE_DRIVE_FOLDER_ID:
        drive_url = f"https://drive.google.com/drive/folders/{GOOGLE_DRIVE_FOLDER_ID}"
        st.sidebar.link_button(t("open_drive"), drive_url)

    service = get_drive_service()
    if service:
        st.sidebar.success(t("connected"))
        if st.sidebar.button(t("disconnect")):
            if os.path.exists(TOKEN_FILE):
                os.remove(TOKEN_FILE)
            st.rerun()
    else:
        st.sidebar.warning(t("not_connected"))
        if os.path.exists(CREDENTIALS_FILE):
            flow = InstalledAppFlow.from_client_secrets_file(
                CREDENTIALS_FILE, SCOPES, redirect_uri='urn:ietf:wg:oauth:2.0:oob'
            )
            auth_url, _ = flow.authorization_url(prompt='consent')
            st.sidebar.markdown(f"{t('auth_step1')}({auth_url})**")
            auth_code = st.sidebar.text_input(t("auth_step2"))
            if st.sidebar.button(t("auth_step3")):
                if auth_code:
                    try:
                        flow.fetch_token(code=auth_code)
                        creds = flow.credentials
                        with open(TOKEN_FILE, 'w') as token_file:
                            token_file.write(creds.to_json())
                        st.success(t("auth_ok"))
                        st.rerun()
                    except Exception as e:
                        st.sidebar.error(f"{t('auth_fail')} {e}")
                else:
                    st.sidebar.error(t("auth_no_code"))
        else:
            st.sidebar.error(t("no_credentials"))

    st.sidebar.markdown("---")

    st.sidebar.subheader(t("data_mgmt"))
    if 'confirm_delete_history' not in st.session_state:
        st.session_state.confirm_delete_history = False

    if st.sidebar.button(t("delete_history")):
        st.session_state.confirm_delete_history = True

    if st.session_state.confirm_delete_history:
        st.sidebar.error(t("confirm_delete"))
        col_del_1, col_del_2 = st.sidebar.columns(2)
        if col_del_1.button(t("yes_delete")):
            for p in [LOCAL_HISTORY_PATH, LOCAL_TEST_HIST_PATH]:
                if os.path.exists(p):
                    os.remove(p)
            st.session_state.confirm_delete_history = False
            st.rerun()
        if col_del_2.button(t("cancel")):
            st.session_state.confirm_delete_history = False
            st.rerun()

    # -----------------------------------------------------------------------
    # Klienta dati
    # -----------------------------------------------------------------------

    if 'client_data' not in st.session_state:
        st.session_state.client_data = {'name': '', 'address': '', 'reg_no': '', 'vat_no': ''}
    if 'e_invoice_data' not in st.session_state:
        st.session_state.e_invoice_data = {
            'receiver_name': '', 'receiver_reg_no': '', 'receiver_address': '',
            'customer_name': '', 'customer_reg_no': '', 'customer_address': ''
        }

    # Map doc_type to internal LV key for logic
    lang = st.session_state.get("lang", "lv")
    lv_doc_types = TRANSLATIONS["lv"]["doc_types"]
    en_doc_types = TRANSLATIONS["en"]["doc_types"]
    if lang == "en" and doc_type in en_doc_types:
        doc_type_lv = lv_doc_types[en_doc_types.index(doc_type)]
    else:
        doc_type_lv = doc_type

    if doc_type_lv != "E-rēķins":
        st.header(t("client"))
        col1, col2 = st.columns([1, 1])
        with col1:
            lursoft_url = st.text_input(t("lursoft_url"))
            scrape_btn  = st.button(t("load_lursoft"))
            if scrape_btn and lursoft_url:
                with st.spinner(t("loading")):
                    scraped = scrape_lursoft(lursoft_url)
                    if scraped:
                        if scraped.get('name'):    st.session_state.client_data['name']    = scraped['name']
                        if scraped.get('address'): st.session_state.client_data['address'] = scraped['address']
                        if scraped.get('reg_no'):  st.session_state.client_data['reg_no']  = scraped['reg_no']
                        st.session_state.client_data['vat_no'] = "LV" + scraped.get('reg_no', '')
                        st.success(t("lursoft_ok"))
                        st.rerun()
                    else:
                        st.error(t("lursoft_fail"))
        with col2:
            st.session_state.client_data['name']    = st.text_input(t("name"),    value=st.session_state.client_data['name'])
            st.session_state.client_data['address'] = st.text_input(t("address"), value=st.session_state.client_data['address'])
            st.session_state.client_data['reg_no']  = st.text_input(t("reg_no"),  value=st.session_state.client_data['reg_no'])
            st.session_state.client_data['vat_no']  = st.text_input(t("vat_no"),  value=st.session_state.client_data['vat_no'])
    else:
        st.header(f"{t('receiver')} & {t('customer')}")
        col_rec, col_cus = st.columns(2)
        with col_rec:
            st.subheader(t("receiver"))
            rec_name = st.text_input(t("receiver_inst"), value=st.session_state.e_invoice_data.get('receiver_name', ''), key='rec_n')
            rec_reg  = st.text_input(t("receiver_reg"),  value=st.session_state.e_invoice_data.get('receiver_reg_no', ''), key='rec_r')
            rec_addr = st.text_input(t("receiver_addr"), value=st.session_state.e_invoice_data.get('receiver_address', ''), key='rec_a')
            st.session_state.e_invoice_data['receiver_name']    = rec_name
            st.session_state.e_invoice_data['receiver_reg_no']  = rec_reg
            st.session_state.e_invoice_data['receiver_address'] = rec_addr
        with col_cus:
            st.subheader(t("customer"))
            cus_name = st.text_input(t("customer_name"), value=st.session_state.e_invoice_data.get('customer_name', ''), key='cus_n')
            cus_reg  = st.text_input(t("customer_reg"),  value=st.session_state.e_invoice_data.get('customer_reg_no', ''), key='cus_r')
            cus_addr = st.text_input(t("customer_addr"), value=st.session_state.e_invoice_data.get('customer_address', ''), key='cus_a')
            st.session_state.e_invoice_data['customer_name']    = cus_name
            st.session_state.e_invoice_data['customer_reg_no']  = cus_reg
            st.session_state.e_invoice_data['customer_address'] = cus_addr

    # -----------------------------------------------------------------------
    # Preces / Pakalpojumi
    # -----------------------------------------------------------------------

    st.markdown("---")
    st.header(t("goods"))

    if 'items_df' not in st.session_state:
        st.session_state.items_df = pd.DataFrame(
            columns=["NOSAUKUMS", "Mērvienība", "DAUDZUMS", "CENA (EUR)"]
        )

    st.subheader(t("add_from_presets"))
    presets_df = load_presets()
    if not presets_df.empty:
        p_col1, p_col2, p_col3 = st.columns([3, 1, 1])
        with p_col1:
            preset_options  = presets_df['NOSAUKUMS'].tolist()
            selected_preset = st.selectbox(t("choose_product"), preset_options)
        with p_col2:
            preset_qty = st.number_input(t("quantity"), min_value=0.01, value=1.00, step=0.01)
        with p_col3:
            st.write("")
            st.write("")
            if st.button(t("add_to_table")):
                sel_row  = presets_df[presets_df['NOSAUKUMS'] == selected_preset].iloc[0]
                new_item = {
                    "NOSAUKUMS":  sel_row['NOSAUKUMS'],
                    "Mērvienība": sel_row['Mērvienība'],
                    "DAUDZUMS":   preset_qty,
                    "CENA (EUR)": sel_row['CENA (EUR)']
                }
                st.session_state.items_df = pd.concat(
                    [st.session_state.items_df, pd.DataFrame([new_item])], ignore_index=True
                )
                st.rerun()
    else:
        st.info(t("presets_empty"))

    display_df = st.session_state.items_df.copy()
    display_df['DAUDZUMS']        = pd.to_numeric(display_df['DAUDZUMS'],        errors='coerce').fillna(0)
    display_df['CENA (EUR)']      = pd.to_numeric(display_df['CENA (EUR)'],      errors='coerce').fillna(0)
    display_df['Cena kopā (EUR)'] = display_df['DAUDZUMS'] * display_df['CENA (EUR)']

    edited_df = st.data_editor(
        display_df, num_rows="dynamic", width="stretch", hide_index=False,
        column_config={
            "CENA (EUR)":      st.column_config.NumberColumn(format="%.2f"),
            "DAUDZUMS":        st.column_config.NumberColumn(format="%.2f", step=0.01),
            "Cena kopā (EUR)": st.column_config.NumberColumn("Cena kopā (EUR)", disabled=True, format="%.2f")
        }
    )

    if st.button(t("recalculate")):
        st.session_state.items_df = edited_df.drop(columns=['Cena kopā (EUR)'], errors='ignore')
        st.rerun()

    # -----------------------------------------------------------------------
    # Aprēķini
    # -----------------------------------------------------------------------

    subtotal, vat, total             = 0.0, 0.0, 0.0
    advance_payment, advance_percent = 0.0, 0.0
    discount_eur, discount_percent   = 0.0, 0.0
    subtotal_after_discount          = 0.0
    amount_words                     = ""
    apply_vat                        = False
    calc_df                          = edited_df.copy()

    def fmt_curr(val):
        return f"{val:,.2f}".replace(",", "X").replace(".", ",").replace("X", " ")

    try:
        if not edited_df.empty:
            calc_df['DAUDZUMS']   = pd.to_numeric(calc_df['DAUDZUMS'],   errors='coerce').fillna(0)
            calc_df['CENA (EUR)'] = pd.to_numeric(calc_df['CENA (EUR)'], errors='coerce').fillna(0)
            calc_df['KOPĀ (EUR)'] = calc_df['DAUDZUMS'] * calc_df['CENA (EUR)']
            subtotal = calc_df['KOPĀ (EUR)'].sum()

            st.markdown(t("discount"))
            discount_options = [t("no_discount"), t("percent_discount"), t("eur_discount")]
            discount_type = st.radio(t("discount_type"), discount_options, horizontal=True)
            if discount_type == t("percent_discount"):
                discount_percent = st.number_input(t("discount_pct"), 0.0, 100.0, 0.0, 5.0)
                discount_eur     = subtotal * (discount_percent / 100)
            elif discount_type == t("eur_discount"):
                discount_eur     = st.number_input(t("discount_eur_input"), 0.0, subtotal, 0.0, 10.0)
                discount_percent = (discount_eur / subtotal * 100) if subtotal > 0 else 0

            subtotal_after_discount = subtotal - discount_eur

            # PVN ķeksītis — noklusējums: NEATZĪMĒTS
            apply_vat = st.checkbox(t("vat_toggle"), value=False, key="apply_vat_checkbox")

            vat   = subtotal_after_discount * 0.21 if apply_vat else 0.0
            total = subtotal_after_discount + vat

            is_advance = "avansa" in doc_type_lv.lower() or "advance" in doc_type.lower()

            if is_advance:
                st.markdown(t("advance_settings"))
                calc_options = [t("in_eur"), t("in_pct")]
                calc_method = st.radio(t("calc_method"), calc_options, horizontal=True)
                if calc_method == t("in_eur"):
                    advance_payment = st.number_input(t("in_eur"), 0.0, total, total, 10.0)
                    advance_percent = (advance_payment / total * 100) if total > 0 else 0
                else:
                    advance_percent = st.number_input(t("in_pct"), 0.0, 100.0, 50.0, 5.0)
                    advance_payment = total * (advance_percent / 100)
                _, t_col2 = st.columns([3, 1])
                with t_col2:
                    st.markdown(f"{t('total_order')} € {fmt_curr(total)}")
                    st.markdown(f"**{t('advance_payable')} ({int(round(advance_percent))}%):** € {fmt_curr(advance_payment)}")
                amount_words = money_to_words_lv(advance_payment)
                st.info(f"{t('amount_words_label')} {amount_words}")
            else:
                advance_payment = total
                _, t_col2 = st.columns([3, 1])
                with t_col2:
                    st.markdown(f"{t('total_no_vat')} € {fmt_curr(subtotal)}")
                    if discount_eur > 0:
                        st.markdown(f"{t('discount_amount')} ({discount_percent:g}%)**:** € -{fmt_curr(discount_eur)}")
                        st.markdown(f"{t('total_with_discount')} € {fmt_curr(subtotal_after_discount)}")
                    st.markdown(f"{t('vat_21')} € {fmt_curr(vat)}")
                    st.markdown(f"{t('total_payable')} € {fmt_curr(total)}")
                amount_words = money_to_words_lv(total)
                st.info(f"{t('amount_words_full')} {amount_words}")
    except Exception as e:
        st.error(f"{t('calc_error')} {e}")

    # -----------------------------------------------------------------------
    # Komentāri un Paraksti
    # -----------------------------------------------------------------------

    st.markdown("---")
    st.header(t("comments_header"))

    comments = st.text_area(
        t("comments_label"),
        value=st.session_state.get('loaded_comments', '')
    )

    col_sig1, col_sig2 = st.columns(2)
    with col_sig1:
        signatory_title = st.text_input(t("title"), t("default_title"), key="sig_title")
    with col_sig2:
        st.write("")
    full_signatory = f"SIA Baltic SEO {signatory_title} Adrians Stankevičs"
    st.caption(f"{t('signatory_preview')} {full_signatory}")

    # -----------------------------------------------------------------------
    # invoice_data
    # -----------------------------------------------------------------------

    invoice_data = {
        'doc_type':                doc_type_lv,
        'doc_id':                  doc_id,
        'date':                    doc_date.strftime("%d.%m.%Y"),
        'due_date':                due_date.strftime("%d.%m.%Y"),
        'client_name':             st.session_state.client_data['name'],
        'client_address':          st.session_state.client_data['address'],
        'client_reg_no':           st.session_state.client_data['reg_no'],
        'client_vat_no':           st.session_state.client_data['vat_no'],
        'items':                   [],
        'subtotal':                fmt_curr(subtotal),
        'vat':                     fmt_curr(vat),
        'total':                   fmt_curr(total),
        'raw_total':               total,
        'apply_vat':               apply_vat,
        'raw_advance':             advance_payment,
        'advance_percent':         advance_percent,
        'discount_eur':            fmt_curr(discount_eur),
        'raw_discount_eur':        discount_eur,
        'discount_percent':        discount_percent,
        'subtotal_after_discount': fmt_curr(subtotal_after_discount),
        'amount_words':            amount_words,
        'signatory':               full_signatory,
        'comments':                comments,
        'lang':                    lang,
        'receiver_name':           st.session_state.get('e_invoice_data', {}).get('receiver_name', ''),
        'receiver_reg_no':         st.session_state.get('e_invoice_data', {}).get('receiver_reg_no', ''),
        'receiver_address':        st.session_state.get('e_invoice_data', {}).get('receiver_address', ''),
        'customer_name':           st.session_state.get('e_invoice_data', {}).get('customer_name', ''),
        'customer_reg_no':         st.session_state.get('e_invoice_data', {}).get('customer_reg_no', ''),
        'customer_address':        st.session_state.get('e_invoice_data', {}).get('customer_address', ''),
    }

    if not edited_df.empty:
        for _, row in calc_df.iterrows():
            invoice_data['items'].append({
                'seq':       len(invoice_data['items']) + 1,
                'name':      row.get('NOSAUKUMS', ''),
                'unit':      row.get('Mērvienība', ''),
                'qty':       str(row.get('DAUDZUMS', 0)),
                'price':     fmt_curr(row.get('CENA (EUR)', 0)),
                'total':     fmt_curr(row.get('KOPĀ (EUR)', 0)),
                'raw_qty':   float(row.get('DAUDZUMS', 0)),
                'raw_price': float(row.get('CENA (EUR)', 0))
            })

    # -----------------------------------------------------------------------
    # Lejupielāde
    # -----------------------------------------------------------------------

    st.markdown("---")
    st.markdown(t("download_header"))

    is_proforma = st.toggle(
        t("proforma_toggle"), value=False, help=t("proforma_help")
    )

    if is_proforma:
        proforma_map = t("proforma_map")
        invoice_data['doc_type'] = proforma_map.get(doc_type_lv, doc_type_lv)

    d_col1, d_col2 = st.columns(2)

    try:
        pdf_file      = generate_pdf(invoice_data)
        file_name_pdf = f"{invoice_data['doc_type'].replace(' ', '_')}_{doc_id.replace(' ', '_')}.pdf"
        with d_col1:
            st.download_button(
                label=t("download_pdf"),
                data=pdf_file,
                file_name=file_name_pdf,
                mime="application/pdf",
                on_click=handle_download,
                args=(invoice_data, pdf_file, file_name_pdf, "application/pdf", is_proforma)
            )
    except Exception as e:
        st.error(f"{t('pdf_error')} {e}")

    try:
        docx_file      = generate_docx(invoice_data)
        file_name_docx = f"{invoice_data['doc_type'].replace(' ', '_')}_{doc_id.replace(' ', '_')}.docx"
        with d_col2:
            st.download_button(
                label=t("download_word"),
                data=docx_file,
                file_name=file_name_docx,
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                on_click=handle_download,
                args=(invoice_data, docx_file, file_name_docx,
                      "application/vnd.openxmlformats-officedocument.wordprocessingml.document", is_proforma)
            )
    except Exception as e:
        st.error(f"{t('word_error')} {e}")

    # -----------------------------------------------------------------------
    # Vēstures tabula
    # -----------------------------------------------------------------------

    st.markdown("---")
    with st.expander(t("history_expander"), expanded=False):
        if history:
            excel_bytes = generate_history_excel(history)
            st.download_button(
                label=t("download_excel"),
                data=excel_bytes,
                file_name=f"rekinu_vesture_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.markdown("---")

            rows_html = ""
            for entry in history:
                rows_html += f"""
                <tr>
                    <td style="text-align:center">{entry.get('kartas_nr', '')}</td>
                    <td style="text-align:center">{entry.get('datums', entry.get('date', ''))}</td>
                    <td>{entry.get('pr_partneris', entry.get('client_name', ''))}</td>
                    <td style="text-align:center">{entry.get('pr_pvn_nr', entry.get('client_vat_no', ''))}</td>
                    <td style="text-align:center">{entry.get('pr_datums', entry.get('date', ''))}</td>
                    <td style="text-align:center">{entry.get('pr_numurs', entry.get('doc_id', ''))}</td>
                    <td style="max-width:220px; white-space:normal; word-break:break-word">{entry.get('darijuma_apraksts', '')}</td>
                    <td style="text-align:right">{entry.get('vertiba_bez_pvn', '')}</td>
                    <td style="text-align:center">{entry.get('dabas_resursi', '') or '—'}</td>
                    <td style="text-align:right">{entry.get('atlaides', '') or '—'}</td>
                    <td style="text-align:right">{entry.get('pvn_summa', '')}</td>
                    <td style="text-align:right; font-weight:bold">{entry.get('kopeja_summa', entry.get('total', ''))}</td>
                </tr>"""

            st.markdown(f"""
            <style>
            .inv-hist {{
                border-collapse: collapse;
                font-size: 11px;
                width: 100%;
            }}
            .inv-hist th, .inv-hist td {{
                border: 1px solid #4da6e8;
                padding: 4px 7px;
                vertical-align: middle;
            }}
            .inv-hist thead th {{
                background-color: #1a6fa8;
                color: white;
                text-align: center;
                font-weight: bold;
                line-height: 1.3;
            }}
            .inv-hist tbody tr:nth-child(even) {{
                background-color: #eaf4fb;
            }}
            .inv-hist tbody tr:hover {{
                background-color: #c2e0f4;
            }}
            </style>
            <div style="overflow-x:auto; margin-top:10px">
            <table class="inv-hist">
                <thead>
                    <tr>
                        <th rowspan="2" style="min-width:50px">Kārtas<br>Nr.</th>
                        <th rowspan="2" style="min-width:80px">Datums</th>
                        <th rowspan="2" style="min-width:150px">PR norādītais<br>darījuma partneris</th>
                        <th rowspan="2" style="min-width:140px">PR norādītā darījuma<br>partnera reģistrācijas<br>vai PVN maksātāja Nr.</th>
                        <th colspan="2">PR datums un numurs</th>
                        <th rowspan="2" style="min-width:180px">Darījuma apraksts</th>
                        <th rowspan="2" style="min-width:100px">PR norādītā<br>darījuma vērtība<br>(bez PVN)</th>
                        <th rowspan="2" style="min-width:100px">Dabas resursu<br>un akcīzes<br>nodokļi</th>
                        <th rowspan="2" style="min-width:90px">Piešķirtās<br>atlaides</th>
                        <th rowspan="2" style="min-width:90px">PVN summa</th>
                        <th rowspan="2" style="min-width:100px">Kopējā<br>summa</th>
                    </tr>
                    <tr>
                        <th style="min-width:80px">Datums</th>
                        <th style="min-width:90px">Numurs</th>
                    </tr>
                </thead>
                <tbody>{rows_html}</tbody>
            </table>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.info(t("history_empty"))

# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main():
    if "lang" not in st.session_state:
        st.session_state.lang = "lv"

    st.title(t("app_title"))

    tab_invoice, tab_presets = st.tabs([t("tab_invoice"), t("tab_presets")])
    with tab_invoice:
        render_invoice_app()
    with tab_presets:
        render_presets_app()

if __name__ == "__main__":
    main()
